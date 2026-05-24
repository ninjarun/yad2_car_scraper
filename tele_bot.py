#!/usr/bin/env python3
"""
tele_bot.py — Telegram UI wrapper for yad2.py

Goal: Let users pick brands & models in Telegram *without modifying* your existing
scraper (yad2.py). We achieve this by:
  1) Loading the same value4urlBuild.json to build Telegram inline keyboards.
  2) Converting the user's selections into the exact terminal inputs that yad2.py
     expects (first a comma‑separated list of 1‑based brand indices, then for each
     selected brand a comma‑separated list of 1‑based model indices on its own line).
  3) Launching yad2.py as a subprocess and piping those inputs to its stdin.
  4) Printing scraper logs to your local terminal (STDOUT) only.
  5) After completion, sending a Telegram summary filtered by the user's selection
     and attaching the full CSV.

Usage:
  export TELEGRAM_TOKEN=xxxx:yyyy
  python3 tele_bot.py
  

Dependencies (Python 3.10+):
  pip install python-telegram-bot==21.4

Files expected in the same folder:
  - yad2.py
  - value4urlBuild.json
  - (optional) yad2_cars_data.csv — created/updated by yad2.py
"""


from __future__ import annotations

import os
import json
import math
import textwrap
import asyncio
import subprocess
import time
import logging
import csv
import sqlite3
from pathlib import Path
from typing import Dict, List, Set, Optional, Tuple
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from collections import defaultdict

from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    LinkPreviewOptions,
)
from telegram.constants import ParseMode
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes,
    Defaults,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment
import re
from db import (
    export_csv,
    query_selection,
    init_subscriptions,
    upsert_user,
    add_subscription,
    list_subscriptions,
    list_user_subscriptions,
    diff_and_update_items,
    update_subscription_checked,
    init_job_runs,
    recent_job_exists,
    record_job_run,
    reset_all_subscriptions,
    get_plan_limit,
    list_free_trial_expired_users,
    reset_user_subscriptions,
    init_plans,
    is_free_plan,
    mark_free_reset,
    set_user_plan,
    get_plan
)
from tele_bot_stripe import register_stripe_handlers, boot_stripe_webhook_once
from datetime import datetime, timezone
from telegram.constants import ParseMode
from telegram import InlineKeyboardMarkup, InlineKeyboardButton

############################################################################################
#  Admin Functions
############################################################################################
# --- Admin allow-list ---
ADMINS = {366607102}  # <-- your Telegram user id here (add others as needed)

def _is_admin(user_id: int) -> bool:
    return user_id in ADMINS
async def myid_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Let any user see their own Telegram ID."""
    uid = update.effective_user.id
    await update.message.reply_text(
        f"🔎 ה־ID שלך הוא:\n`{uid}`",
        parse_mode=ParseMode.MARKDOWN,
    )

async def get_plan_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /get_plan            -> show caller's plan
    /get_plan <tg_id>    -> (admin only) show plan for another user
    """
    caller_id = update.effective_user.id
    text = (update.message.text or "").strip()
    parts = text.split()

    target_id = caller_id

    # If an ID is provided, only admins are allowed to inspect others
    if len(parts) >= 2:
        if not _is_admin(caller_id):
            await update.message.reply_text("⛔ רק אדמין יכול לבדוק משתמש אחר.")
            return
        try:
            target_id = int(parts[1])
        except ValueError:
            await update.message.reply_text("❌ tg_id לא תקין.")
            return

    # Get plan + limit (this already enforces expiry / free window)
    plan_name, max_subs = get_plan_limit(target_id)

    # Get raw plan row to show expiry (if any)
    plan_row = get_plan(target_id)
    expires_at = plan_row["expires_at"] if plan_row and plan_row.get("expires_at") else None

    if target_id == caller_id:
        header = "📦 פרטי המסלול שלך:"
    else:
        header = f"📦 פרטי המסלול עבור משתמש {target_id}:"

    if expires_at:
        exp_human = _format_dt(expires_at)   # existing helper in this file
        exp_line = f"⏱ תוקף עד: {exp_human}"
    else:
        exp_line = "⏱ תוקף: ללא תאריך תפוגה (מסלול חינמי או ללא הגבלה)."

    msg = (
        f"{header}\n\n"
        f"🔖 מסלול: {plan_name}\n"
        f"📌 מקסימום מנויים: {max_subs}\n"
        f"{exp_line}"
    )

    await update.message.reply_text(msg)


async def set_plan_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Admin-only:
    /set_plan <tg_id> <plan_name> <months|lifetime>
    Example: /set_plan 366607102 pro 3
             /set_plan 366607102 dealer lifetime
    """
    # Admin check
    uid = update.effective_user.id
    if not _is_admin(uid):
        await update.message.reply_text("⛔ Admins only.")
        return

    text = (update.message.text or "").strip()
    parts = text.split()

    if len(parts) < 4:
        await update.message.reply_text(
            "❌ שימוש לא נכון.\n"
            "/set_plan <tg_id> <plan_name> <months|lifetime>\n\n"
            "דוגמה:\n"
            "/set_plan 366607102 pro 3\n"
            "/set_plan 366607102 dealer lifetime"
        )
        return

    # parse tg_id
    try:
        target_tg_id = int(parts[1])
    except ValueError:
        await update.message.reply_text("❌ tg_id לא תקין.")
        return

    plan_name = parts[2].lower()
    duration_raw = parts[3].lower()

    # parse duration
    months: int | None
    if duration_raw == "lifetime":
        months = None
    else:
        try:
            months_int = int(duration_raw)
            if months_int <= 0:
                raise ValueError
            months = months_int
        except ValueError:
            await update.message.reply_text("❌ החודשים לא תקינים. השתמש במספר חיובי או lifetime.")
            return

    # update DB
    try:
        set_user_plan(target_tg_id, plan_name, months)
    except Exception as e:
        await update.message.reply_text(f"❌ שגיאה בעדכון מסד הנתונים: {e}")
        return

    if months is None:
        duration_txt = "ללא הגבלת זמן (lifetime)"
    else:
        duration_txt = f"{months} חודשים"

    await update.message.reply_text(
        "✅ המסלול עודכן בהצלחה.\n\n"
        f"👤 משתמש: {target_tg_id}\n"
        f"📦 מסלול: {plan_name}\n"
        f"⏱ משך: {duration_txt}"
    )


async def reset_free_subs_cmd(update, context):
    # Admin check
    uid = update.effective_user.id
    if not _is_admin(uid):
        await update.message.reply_text("⛔ Admins only.")
        return

    args = (update.message.text or "").split()
    if len(args) < 2 or args[1].lower() != "confirm":
        await update.message.reply_text(
            "⚠️ This will DELETE subscriptions only for users on the FREE plan.\n\n"
            "Type:\n`/reset_free_subs confirm`\n\nto proceed.",
            parse_mode=ParseMode.MARKDOWN
        )
        return

    expired = list_free_trial_expired_users(days=10)   # [(tg_id, started_iso), ...]
    total_deleted = 0
    for tg_id, started_iso in expired:
        deleted = reset_user_subscriptions(tg_id)
        total_deleted += deleted
        try:
            kb = InlineKeyboardMarkup([
                # [InlineKeyboardButton("💳 הרשמה למסלול", callback_data="register_plan")],
                [InlineKeyboardButton("🏠 תפריט ראשי",   callback_data="home")],
            ])
            await update.get_bot().send_message(
                chat_id=tg_id,
                text=(
                    "🔔 תקופת הניסיון בחינם (10 ימים) הסתיימה, לכן איפסתי את המנויים שלך ל-0.\n"
                    "מעוניין/ת להמשיך לקבל התראות? אפשר לשדרג למסלול בתשלום:"
                ),
                reply_markup=kb,
                disable_web_page_preview=True,
            )
        except Exception:
            pass

    await update.message.reply_text(f"🧹 Done. Cleared {total_deleted} subscriptions for FREE users.")


async def reset_all_subs_cmd(update, context):
    # Admin check
    uid = update.effective_user.id
    if not _is_admin(uid):
        await update.message.reply_text("⛔ Admins only.")
        return

    # Require explicit confirmation text to prevent accidents
    args = (update.message.text or "").split()
    if len(args) < 2 or args[1].lower() != "confirm":
        await update.message.reply_text(
            "⚠️ This will DELETE *all* users' subscriptions.\n\n"
            "Type:\n`/reset_all_subs confirm`\n\nto proceed.",
            parse_mode=ParseMode.MARKDOWN
        )
        return

    subs_deleted, items_deleted = reset_all_subscriptions()
    await update.message.reply_text(
        f"🧹 Done. Deleted {subs_deleted} subscriptions and {items_deleted} subscription_items."
    )


############################################################################################
# ----------------------------- Logging & queue status -----------------------------
############################################################################################


# 1. Base logging level: WARNING globally so we don't get spam
logging.basicConfig(level=logging.WARNING)

# 2. Mute noisy third-party loggers so console stays readable
for _name in [
    "telegram", "telegram.ext", "httpx", "aiohttp", "urllib3",
    "apscheduler", "undetected_chromedriver", "selenium"
]:
    logging.getLogger(_name).setLevel(logging.ERROR)

# 3. Quiet normal tele_bot logs
logging.getLogger("tele_bot").setLevel(logging.WARNING)

# 4. Queue logger: this is what prints the dashboard you care about
QUEUE_LOG = logging.getLogger("queue_dashboard")
QUEUE_LOG.setLevel(logging.INFO)
QUEUE_LOG.propagate = False  # don't double-print to root

_qh = logging.StreamHandler()
_qh.setFormatter(logging.Formatter("%(asctime)s - %(message)s"))
QUEUE_LOG.addHandler(_qh)

# 5. Logger for "skipped due to last {RUN_WINDOW_MINUTES}m" events
SKIP_LOG = logging.getLogger("job_skips")
SKIP_LOG.setLevel(logging.INFO)
SKIP_LOG.propagate = False
_sh = logging.StreamHandler()
_sh.setFormatter(logging.Formatter("%(asctime)s - %(message)s"))
SKIP_LOG.addHandler(_sh)

###########################################################################################

# Queue logging dedupe / throttle
LAST_QUEUE_SNAPSHOT: tuple | None = None
LAST_QUEUE_TS: float = 0.0
QUEUE_COOLDOWN_SEC: float = 4.0  # minimal interval between identical prints

# ==== Scheduling / dedupe window ====
RUN_WINDOW_MINUTES = 60            # was 20
SCHEDULER_INTERVAL_SECONDS = 60 * 60   # 3600s (every 60m)

# throttle queue dashboard so it won’t spam
_LAST_QUEUE_SNAPSHOT = ""
_LAST_QUEUE_TS = 0.0

# ===== Queue priority policy (scheduled-dominant) =====
# Run ~75% scheduled, ~25% manual, with starvation protection for manual.
SCHEDULED_WEIGHT = 3   # how many scheduled in a row
MANUAL_WEIGHT    = 1   # then how many manual in a row
# internal tiny state to count the current streak
_POLICY_STREAK = {"kind": None, "count": 0}


# ===== Queue priority policy (scheduled-dominant) =====
# Default steady-state ratio: ~75% scheduled, ~25% manual
SCHEDULED_WEIGHT = 3
MANUAL_WEIGHT    = 1

# internal streak counter
_POLICY_STREAK = {"kind": None, "count": 0, "weights": (SCHEDULED_WEIGHT, MANUAL_WEIGHT)}

# ----- Light aging (manual burst protection) -----
# If manual backlog gets big, temporarily relax policy to 2:1 (scheduled:manual)
AGING_MANUAL_ON_THRESHOLD  = 12   # when MANUAL_QUEUE >= this -> turn aging ON
AGING_MANUAL_OFF_THRESHOLD = 6   # when MANUAL_QUEUE <= this -> turn aging OFF (hysteresis)
AGING_SCHEDULED_WEIGHT     = 2
AGING_MANUAL_WEIGHT        = 1
_AGING_ACTIVE = False

# ==== TEST MODE (set QUEUE_TEST=1 in env to enable) ====
TEST_MODE = os.environ.get("QUEUE_TEST") == "1"


def _job_label(job: "ScrapeJob") -> str:
    # Compact, info-rich label for console
    src = "🧑‍💻 manual" if getattr(job, "source", "") == "manual" else "⏰ scheduled"
    # brand summary
    if getattr(job, "brand_names", []):
        b = ", ".join(job.brand_names[:2])
        if len(job.brand_names) > 2:
            b += " +…"
    else:
        b = "—"
    # model summary
    if getattr(job, "model_names", []):
        m = ", ".join(job.model_names[:2])
        if len(job.model_names) > 2:
            m += " +…"
    else:
        m = "ALL"
    return f"{src} • {b} / {m} • uid={getattr(job, 'telegram_id', '—')}"

def _queue_snapshot_tuple(pending_manual, pending_sched, current_label: str | None):
    # snapshot = (current, counts, next5 labels)
    next5 = []
    for j in ([*pending_manual, *pending_sched][:5]):
        next5.append(_job_label(j))
    return (
        current_label or "—",
        len(pending_manual),
        len(pending_sched),
        tuple(next5),
    )

def _fmt_job_label(j: "ScrapeJob") -> str:
    kind = "🧑‍💻 manual" if getattr(j, "source", "") == "manual" else "⏰ scheduled"
    brand = " / ".join(getattr(j, "brand_names", []) or []) or "—"
    model = " / ".join(getattr(j, "model_names", []) or []) or "כל הדגמים"
    uid   = getattr(j, "telegram_id", "?")
    return f"{kind} • {brand} / {model} • uid={uid}"

def print_queue_status(current: str | None = None):
    global _LAST_QUEUE_SNAPSHOT, _LAST_QUEUE_TS

    # snapshot both queues without consuming them
    try:
        pending_manual = list(MANUAL_QUEUE._queue)  # type: ignore[attr-defined]
    except Exception:
        pending_manual = []
    try:
        pending_sched = list(SCHEDULED_QUEUE._queue)  # type: ignore[attr-defined]
    except Exception:
        pending_sched = []

    total = len(pending_manual) + len(pending_sched)
    manual_count = len(pending_manual)
    sched_count  = len(pending_sched)

    # Build preview list (limit to next 5 jobs for display)
    preview_list = [("🧑‍💻", j) for j in pending_manual] + [("⏰", j) for j in pending_sched]
    next_lines = [f"      • {_fmt_job_label(j)}" for _, j in preview_list[:5]]

    # === Build snapshot lines (create 'snap' BEFORE using it) ===
    snap: list[str] = []
    snap.append("📋 Queue")

    # Pull current weights (aging-aware) and show the active policy
    try:
        sched_w, manual_w, aging = _current_policy_weights()
        snap.append(f"   ⚖️  Policy: ⏰ {sched_w}:{manual_w} (scheduled:manual){' — aging ACTIVE' if aging else ''}")
    except NameError:
        # If aging helpers not present yet, fall back to static constants
        snap.append(f"   ⚖️  Policy: ⏰ {SCHEDULED_WEIGHT}:{MANUAL_WEIGHT} (scheduled:manual)")

    snap.append(f"   🔍 Now: {current or '—'}")

    if total:
        snap.append(f"   ⏭️  Next ({total} total — 🧑‍💻 {manual_count} manual, ⏰ {sched_count} scheduled):")
        snap.extend(next_lines)
    else:
        snap.append("   ⏭️  Next: —")

    snapshot = "\n".join(snap)

    # Throttle printing identical snapshots
    now = time.time()
    if snapshot != _LAST_QUEUE_SNAPSHOT or (now - _LAST_QUEUE_TS) > 3.0:
        QUEUE_LOG.info("")
        for line in snapshot.splitlines():
            QUEUE_LOG.info(line)
        QUEUE_LOG.info("")
        _LAST_QUEUE_SNAPSHOT = snapshot
        _LAST_QUEUE_TS = now


def _current_policy_weights() -> tuple[int, int, bool]:
    """
    Decide current (scheduled_weight, manual_weight, aging_active)
    based on MANUAL_QUEUE backlog, with hysteresis.
    """
    global _AGING_ACTIVE

    mqlen = MANUAL_QUEUE.qsize()
    if not _AGING_ACTIVE and mqlen >= AGING_MANUAL_ON_THRESHOLD:
        _AGING_ACTIVE = True
        # reset streak so we adopt the new cadence immediately
        _POLICY_STREAK["kind"] = None
        _POLICY_STREAK["count"] = 0
    elif _AGING_ACTIVE and mqlen <= AGING_MANUAL_OFF_THRESHOLD:
        _AGING_ACTIVE = False
        _POLICY_STREAK["kind"] = None
        _POLICY_STREAK["count"] = 0

    if _AGING_ACTIVE:
        return AGING_SCHEDULED_WEIGHT, AGING_MANUAL_WEIGHT, True
    else:
        return SCHEDULED_WEIGHT, MANUAL_WEIGHT, False




#############################################################################################


# from db import export_csv, query_selection  # NEW
DB_FILENAME = "yad2_cars.db"               # NEW
CSV_FILENAME = "yad2_cars_data.csv"        # keep, but it will be generated on demand


# -----------------------------
# Configuration
# -----------------------------
BRANDS_PER_PAGE = 12
MODELS_PER_PAGE = 12
SCRIPT_FILENAME = "yad2.py"
JSON_FILENAME = "value4urlBuild.json"



# -----------------------------
# Helpers to build keyboards
# -----import sqlite3

def count_user_subscriptions(telegram_id: int) -> int:
    conn = _db()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT COUNT(*)
        FROM subscriptions
        WHERE telegram_id = ?
        """,
        (telegram_id,),
    )
    row = cur.fetchone()
    conn.close()
    return int(row[0] if row and row[0] is not None else 0)


def _db():
    return sqlite3.connect(DB_FILENAME)

def get_user_subscriptions(telegram_id: int) -> list[tuple[int, str, str]]:
    """
    Return user's subscriptions.
    Each row = (row_id, brand, model)
    """
    conn = _db()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT rowid, brand, model
        FROM subscriptions
        WHERE telegram_id = ?
        ORDER BY brand, model
        """,
        (telegram_id,),
    )
    rows = cur.fetchall()
    conn.close()
    return rows

def delete_subscription_row(row_id: int, telegram_id: int) -> None:
    """
    Delete a single subscription by rowid, but only if it belongs to this user.
    """
    conn = _db()
    cur = conn.cursor()
    cur.execute(
        """
        DELETE FROM subscriptions
        WHERE rowid = ?
          AND telegram_id = ?
        """,
        (row_id, telegram_id),
    )
    conn.commit()
    conn.close()

def delete_all_subscriptions(telegram_id: int) -> None:
    """
    Delete all subscriptions for this user.
    """
    conn = _db()
    cur = conn.cursor()
    cur.execute(
        """
        DELETE FROM subscriptions
        WHERE telegram_id = ?
        """,
        (telegram_id,),
    )
    conn.commit()
    conn.close()

async def reset_free_subs_cmd(update, context):
    tg_id = update.effective_user.id
    if not is_free_plan(tg_id):
        await update.message.reply_text("הפקודה זמינה רק במסלול חינמי. רוצה לשדרג?")
        # attach your plans keyboard here
        return

    # show confirm keyboard
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("אישור", callback_data="reset_free_subs:confirm"),
         InlineKeyboardButton("ביטול", callback_data="reset_free_subs:cancel")]
    ])
    await update.message.reply_text("לאפס את כל המנויים שלך?", reply_markup=kb)

async def reset_free_subs_cb(update, context):
    q = update.callback_query
    await q.answer()
    tg_id = q.from_user.id
    action = q.data.split(":")[1]

    if action == "confirm":
        delete_all_subscriptions(tg_id)
        mark_free_reset(tg_id)
        await q.edit_message_text(
            "✅ אופסו כל המנויים שלך. חלון חינמי חדש של 10 ימים התחיל.\n"
            "רוצה ללא מגבלה? הצטרף/פי לאחד המסלולים בתשלום."
        )
        # Optionally send your plans keyboard in a new message
    else:
        await q.edit_message_text("בוטל.")


# async def send_existing_results_to_user(query, brand_names: list[str], model_names: list[str], max_items: int = 20):
async def send_existing_results_to_user(query, brand_names: list[str], model_names: list[str], max_items: int = 20, seller_filter: set[str] | None = None):
    """
    Pulls current rows from SQLite for the selection and sends them as rich cards.
    No scraping is performed.
    """
    # rows = query_selection(brand_names, model_names, limit=300)
    rows = query_selection(brand_names, model_names, limit=300)
    rows = _apply_seller_filter(rows, seller_filter)
    if not rows:
        await query.message.reply_text(
            "אין נתונים קיימים להצגה כרגע.",
            link_preview_options=LinkPreviewOptions(is_disabled=True),
        )
        return

    sent = 0
    for r in rows[:max_items]:
        card = format_listing_card(r)
        await query.message.reply_text(
            card[:4000],
            parse_mode=ParseMode.MARKDOWN,
            link_preview_options=LinkPreviewOptions(is_disabled=True),
        )
        sent += 1

    remaining = max(0, len(rows) - sent)
    if remaining > 0:
        await query.message.reply_text(
            f"…ויש עוד {remaining} תוצאות קיימות במסד הנתונים.",
            link_preview_options=LinkPreviewOptions(is_disabled=True),
        )

    # Attach CSV exported from current DB (no new scrape)
    # await try_send_csv(query)
    # await query.message.reply_text("פעולות נוספות:", reply_markup=_home_keyboard())
    # Attach XLSX of exactly these rows (same styling as Start-menu)
    xlsx_path = _rows_to_styled_xlsx(rows, query.from_user.id)
    brand_str = " + ".join(brand_names) if brand_names else "—"
    # model_names may include None/"" for brand-only; filter empties
    clean_models = [m for m in (model_names or []) if m]
    model_str = " + ".join(clean_models) if clean_models else "כל הדגמים"

    await query.message.reply_document(
        document=open(xlsx_path, "rb"),
        filename=xlsx_path,
        caption=f"תוצאות הסריקה — {brand_str} / {model_str}",
    )
    try:
        os.remove(xlsx_path)
    except Exception:
        pass
    await query.message.reply_text("פעולות נוספות:", reply_markup=_home_keyboard())

def _job_keys(brand_names: list[str], model_names: list[str]) -> tuple[str, str]:
    """
    Normalize brand/model selections into comparable keys for de-dup checks.
    Empty -> "*". Sorting ensures set-equivalence.
    """
    bkey = "|".join(sorted(x.strip() for x in brand_names if x and x.strip())) or "*"
    mkey = "|".join(sorted(x.strip() for x in model_names if x and x.strip())) or "*"
    return bkey, mkey

def chunk_buttons(buttons: List[InlineKeyboardButton], per_row: int = 3) -> List[List[InlineKeyboardButton]]:
    return [buttons[i : i + per_row] for i in range(0, len(buttons), per_row)]

def build_brand_keyboard(
    brands: List[dict],
    selected: Set[int],
    page: int,
) -> InlineKeyboardMarkup:
    total_pages = max(1, math.ceil(len(brands) / BRANDS_PER_PAGE))
    page = max(0, min(page, total_pages - 1))

    start = page * BRANDS_PER_PAGE
    end = start + BRANDS_PER_PAGE

    rows: List[List[InlineKeyboardButton]] = []

    # One brand per row, padded wide
    for idx, brand in enumerate(brands[start:end], start=start):
        mark = "✅ " if idx in selected else ""
        label = _pad_wide(f"{mark}{brand['brand']}", units=12)  # widen the button
        rows.append([
            InlineKeyboardButton(
                label,
                callback_data=f"brand_toggle:{idx}:{page}",
            )
        ])

    # Navigation + done
    nav: List[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️ הקודם", callback_data=f"brand_page:{page-1}"))

    # Finish selection
    nav.append(InlineKeyboardButton("סיימתי", callback_data="brands_done"))

    if page < total_pages - 1:
        nav.append(InlineKeyboardButton("הבא ➡️", callback_data=f"brand_page:{page+1}"))

    if nav:
        rows.append(nav)

    # Optional: Main menu shortcut
    rows.append([InlineKeyboardButton("🏠 תפריט ראשי", callback_data="home")])

    return InlineKeyboardMarkup(rows)



# Wide label padding for inline buttons (non-breaking spaces)
_NBSP = "\u00A0"   # no-break space
_FIG  = "\u2007"   # figure space (width-stable)
def _pad_wide(label: str, units: int = 10) -> str:
    pad = (_NBSP + _FIG) * units
    return f"{pad}{label}{pad}"


def build_model_keyboard(
    brand_idx: int,
    brand: dict,
    selected_models: Set[int],
    page: int,
) -> InlineKeyboardMarkup:
    models = brand.get("models", [])
    total_pages = max(1, math.ceil(len(models) / MODELS_PER_PAGE))
    page = max(0, min(page, total_pages - 1))

    start = page * MODELS_PER_PAGE
    end = start + MODELS_PER_PAGE

    btns = []
    for midx, model in enumerate(models[start:end], start=start):
        mark = "✅ " if midx in selected_models else ""
        label = model.get("model_name") or str(model.get("model_value"))
        btns.append(
            InlineKeyboardButton(
                f"{mark}{label}",
                callback_data=f"model_toggle:{brand_idx}:{midx}:{page}",
            )
        )

    rows = chunk_buttons(btns, per_row=1)

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️ הקודם", callback_data=f"model_page:{brand_idx}:{page-1}"))

    # סיום מותג זה
    nav.append(
        InlineKeyboardButton(
            "סיימתי",
            callback_data=f"models_done:{brand_idx}"
        )
    )

    if page < total_pages - 1:
        nav.append(InlineKeyboardButton("הבא ➡️", callback_data=f"model_page:{brand_idx}:{page+1}"))

    rows.append(nav)
    rows.append([InlineKeyboardButton("🏠 תפריט ראשי", callback_data="home")])

    return InlineKeyboardMarkup(rows)


from telegram import InlineKeyboardMarkup, InlineKeyboardButton

def build_subscriptions_keyboard(sub_rows: list[tuple[int,str,str]]) -> InlineKeyboardMarkup:
    """
    sub_rows = [(rowid, brand, model), ...]
    We'll show each sub with a dedicated remove button.
    At the bottom we'll add 'delete all' if any subs exist.
    """
    rows = []

    for rowid, brand, model in sub_rows:
        if model and model.strip():
            label = f"{brand} — {model}"
        else:
            label = f"{brand} — כל הדגמים"  # "all models" in Hebrew

        rows.append([
            InlineKeyboardButton(
                f"❌ הסר {label}",
                callback_data=f"unsubscribe_one:{rowid}"
            )
        ])

    if sub_rows:
        rows.append([
            InlineKeyboardButton(
                "🗑 מחק את כל המנויים שלי",
                callback_data="unsubscribe_all"
            )
        ])

    return InlineKeyboardMarkup(rows)

async def my_subs(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    subs = get_user_subscriptions(tg_id)

    if not subs:
        await update.message.reply_text("אין לך מנויים כרגע.")  # "You currently have no subscriptions."
        return

    # Build human-readable list in Hebrew for the user:
    lines = []
    for _, brand, model in subs:
        if model and model.strip():
            lines.append(f"- {brand} / {model}")
        else:
            lines.append(f"- {brand} / כל הדגמים")  # all models

    txt = (
        "המנויים הפעילים שלך:\n\n" +
        "\n".join(lines) +
        "\n\nאפשר להסיר מנוי ספציפי או למחוק הכל:"
    )

    await update.message.reply_text(
        txt,
        reply_markup=build_subscriptions_keyboard(subs),
    )

async def unsubscribe_one_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    tg_id = query.from_user.id

    # callback_data format: "unsubscribe_one:<rowid>"
    parts = query.data.split(":")
    if len(parts) != 2:
        await edit_text(query, "שגיאה פנימית (קוד 1).")
        return

    try:
        row_id = int(parts[1])
    except ValueError:
        await edit_text(query, "שגיאה פנימית (קוד 2).")
        return

    # delete the specific subscription
    delete_subscription_row(row_id, tg_id)

    # reload user's remaining subs
    subs = get_user_subscriptions(tg_id)

    if not subs:
        # nothing left, replace message with final text
        await edit_text(
            query,
            "המנוי הוסר. אין לך יותר מנויים פעילים."
        )
        return

    # still have some subs → rebuild the list and keyboard
    lines = []
    for _, brand, model in subs:
        if model and model.strip():
            lines.append(f"- {brand} / {model}")
        else:
            lines.append(f"- {brand} / כל הדגמים")

    txt = (
        "המנוי הוסר.\n\n"
        "המנויים הפעילים שלך עכשיו:\n\n" +
        "\n".join(lines) +
        "\n\nאפשר להסיר עוד:"
    )

    await edit_text(
        query,
        txt,
        reply_markup=build_subscriptions_keyboard(subs),
    )
async def unsubscribe_all_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    tg_id = query.from_user.id

    delete_all_subscriptions(tg_id)

    await edit_text(
        query,
        "כל המנויים שלך נמחקו. לא תקבל יותר התראות.\n\nאפשר תמיד להתחיל מחדש עם /start ולבנות מנויים חדשים."
    )

# -----------------------------
# Bot state per user
# -----------------------------

class Session:
    def __init__(self, brands: List[dict]):
        self.brands = brands
        self.brand_selection: Set[int] = set()
        self.model_selection: Dict[int, Set[int]] = {}
        self.stage = "brands"  # or f"models:{brand_idx}" or "ready"
        self.page_brand = 0
        self.page_model: Dict[int, int] = {}
        self.seller_filter: set[str] = {"private", "agency"}

    def ensure_model_set(self, bidx: int):
        if bidx not in self.model_selection:
            self.model_selection[bidx] = set()

    # Build the exact stdin text yad2.py expects
    def build_stdin_script(self) -> str:
        # IMPORTANT: yad2.py expects 1-based indices, comma-separated
        ordered_brand_indices = sorted(self.brand_selection)
        if not ordered_brand_indices:
            return "\n"  # nothing selected; scraper may exit

        first_line = ",".join(str(i + 1) for i in ordered_brand_indices)

        lines = [first_line]
        for bidx in ordered_brand_indices:
            selected_models = sorted(self.model_selection.get(bidx, []))
            if selected_models:
                line = ",".join(str(m + 1) for m in selected_models)
            else:
                line = ""  # allow empty -> no models for that brand
            lines.append(line)

        return "\n".join(lines) + "\n"

# -----------------------------
# Handlers
# -----------------------------
async def unsubscribe_all_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    tg_id = query.from_user.id

    delete_all_subscriptions(tg_id)

    await edit_text(
        query,
        "כל המנויים שלך נמחקו. לא תקבל יותר התראות.\n\nאפשר תמיד להתחיל מחדש עם /start ולבנות מנויים חדשים.",
    )


async def unsubscribe_one_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    tg_id = query.from_user.id

    # פורמט callback_data = "unsubscribe_one:<rowid>"
    parts = query.data.split(":")
    if len(parts) != 2:
        await edit_text(query, "שגיאה פנימית (קוד 1).")
        return

    try:
        row_id = int(parts[1])
    except ValueError:
        await edit_text(query, "שגיאה פנימית (קוד 2).")
        return

    # מוחק מנוי בודד
    delete_subscription_row(row_id, tg_id)

    # טוען מחדש מנויים ומרענן את ההודעה
    subs = get_user_subscriptions(tg_id)
    if not subs:
        await edit_text(query, "מנוי הוסר. אין לך יותר מנויים פעילים.")
        return

    lines = []
    for _, brand, model in subs:
        if model and model.strip():
            lines.append(f"- {brand} / {model}")
        else:
            lines.append(f"- {brand} / כל הדגמים")

    txt = (
        "המנוי הוסר.\n\n"
        "המנויים הפעילים שלך עכשיו:\n\n" +
        "\n".join(lines) +
        "\n\nאפשר להסיר עוד:"
    )

    await edit_text(
        query,
        txt,
        reply_markup=build_subscriptions_keyboard(subs),
    )


async def my_subs(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    subs = get_user_subscriptions(tg_id)

    if not subs:
        await update.message.reply_text("אין לך מנויים כרגע.")
        return

    # טקסט רשימה לקריאה
    lines = []
    for _, brand, model in subs:
        if model and model.strip():
            lines.append(f"- {brand} / {model}")
        else:
            lines.append(f"- {brand} / כל הדגמים")

    txt = (
        "המנויים הפעילים שלך:\n\n" +
        "\n".join(lines) +
        "\n\nאפשר להסיר מנוי ספציפי או למחוק הכל:"
    )

    await update.message.reply_text(
        txt,
        reply_markup=build_subscriptions_keyboard(subs),
    )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    tg_id = user.id
    username = user.first_name or (user.username or "משתמש")

    # Make sure the user exists in DB
    upsert_user(tg_id, user.username or "")

    # This will NOW call db.get_plan_limit (after the import fix)
    plan_name, max_subs = get_plan_limit(tg_id)

    # Raw plan row so we can see free_started_at / expires_at
    row = get_plan(tg_id) or {}
    free_start = row.get("free_started_at")
    expires_at = row.get("expires_at")

    plan_line = f"• מסלול: {plan_name}\n" f"• מקסימום מנויים: {max_subs}"
    exp_line = ""

    # ----- Free plan: show countdown based on free_started_at -----
    if plan_name.lower() == "free":
        if free_start:
            try:
                # free_started_at is stored as e.g. 2025-11-17T11:21:30Z
                dt_start = datetime.fromisoformat(free_start.replace("Z", "+00:00"))
                dt_exp = dt_start + timedelta(days=10)

                now = datetime.now(dt_start.tzinfo)
                days_left = (dt_exp.date() - now.date()).days

                if days_left <= 0:
                    exp_line = "• ⏱️ תקופת הניסיון החינמית שלך הסתיימה."
                else:
                    exp_line = (
                        "• ⏱️ תקופת הניסיון החינמית שלך "
                        f"מסתיימת ב־{dt_exp.strftime('%d/%m/%Y')} "
                        f"(עוד ~{days_left} ימים)."
                    )
            except Exception:
                # Only if parsing fails for some weird reason
                exp_line = "• ⏱️ תקופת הניסיון החינמית שלך פעילה (עד 10 ימים מהרישום)."
        else:
            # No free_started_at at all → should be rare after the import fix,
            # but keep a safe fallback.
            exp_line = "• ⏱️ תקופת הניסיון החינמית שלך פעילה (עד 10 ימים מהרישום)."

    # ----- Paid plans (pro / dealer / etc.) -----
    else:
        if expires_at:
            # Uses your existing _format_dt helper in this file
            exp_line = f"• ⏱️ המנוי שלך בתוקף עד: {_format_dt(expires_at)}"
        else:
            exp_line = "• ⏱️ המנוי שלך פעיל ללא תאריך תפוגה מוגדר."

    intro_text = (
        f"שלום {username}! 👋\n\n"
        "אני סוכן חכם שסורק את יד2 עבורך 24/7 — אוטומטית וללא תלות בך.\n\n"
        "מה אני עושה עבורך:\n"
        "• עוקב אחר מודעות חדשות שמתפרסמות ביד2.\n"
        "• מזהה ירידות מחיר ושינויים חשובים.\n"
        "• שולח אליך בטלגרם רק דברים חשובים — בזמן אמת.\n"
        "• מייצר עבורך קובץ אקסל מסודר עם כל המנויים והמודעות.\n\n"
        "איך מתחילים:\n"
        "1) בוחרים יצרנים (מותגים) ודגמים שמעניינים אותך.\n"
        "2) אני מטפל בכל השאר ומעדכן אותך אוטומטית לאורך כל היום.\n\n"
        "📦 פרטי המסלול שלך:\n"
        f"{plan_line}\n"
        f"{exp_line}\n\n"
        "ליצירת קשר / תמיכה:\n"
        "@jewishrunner"
    )

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🚀 התחלת סריקה", callback_data="go_scan")],
        [InlineKeyboardButton("📬 המנויים שלי", callback_data="open_my_subs")],
        [InlineKeyboardButton("📥 קובץ אקסל", callback_data="export_subs_csv")],
        # [InlineKeyboardButton("💳 הרשמה למסלול", callback_data="register_plan")],
    ])

    # /start can come from a command or a callback — handle both
    if update.message:
        await update.message.reply_text(
            intro_text,
            reply_markup=kb,
            # parse_mode=ParseMode.MARKDOWN,
        )
        
    else:
        q = update.callback_query
        await q.answer()
        await q.message.reply_text(
            intro_text,
            reply_markup=kb,
            # parse_mode=ParseMode.MARKDOWN,
        )








async def scan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        brands = json.loads(Path(JSON_FILENAME).read_text(encoding="utf-8"))
    except Exception as e:
        await update.message.reply_text(f"לא הצלחתי לטעון {JSON_FILENAME}: {e}")
        return

    context.user_data["session"] = Session(brands)

    await update.message.reply_text("בחר/י מותגים לסריקה (אפשר כמה):")
    await update.message.reply_text(
        "בחירת מותגים:",
        reply_markup=build_brand_keyboard(brands, set(), page=0),
    )




def _get_session(context: ContextTypes.DEFAULT_TYPE) -> Session:
    sess = context.user_data.get("session")
    if not isinstance(sess, Session):
        raise RuntimeError("Session missing. Send /start")
    return sess


async def brand_page_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = _get_session(context)
    page = int(query.data.split(":")[1])
    sess.page_brand = page
    await query.edit_message_reply_markup(
        reply_markup=build_brand_keyboard(sess.brands, sess.brand_selection, page)
    )


async def brand_toggle_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = _get_session(context)
    _, idx_str, page_str = query.data.split(":")
    idx = int(idx_str)
    page = int(page_str)

    if idx in sess.brand_selection:
        sess.brand_selection.remove(idx)
    else:
        sess.brand_selection.add(idx)

    sess.page_brand = page
    await query.edit_message_reply_markup(
        reply_markup=build_brand_keyboard(sess.brands, sess.brand_selection, page)
    )

# ✅ When user taps “Done” on the brands screen
async def brands_done_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = _get_session(context)

    if not sess.brand_selection:
        await edit_text(query, "לא נבחרו מותגים. בחר/י לפחות מותג אחד.")
        return

    # עוברים למותג הראשון שנבחר
    first_brand_idx = sorted(sess.brand_selection)[0]
    sess.stage = f"models:{first_brand_idx}"
    sess.ensure_model_set(first_brand_idx)

    brand = sess.brands[first_brand_idx]
    await edit_text(
        query,
        text=(
            f"בחר/י דגמים עבור {brand['brand']}.\n"
            "כשתסיים/י עם המותג הזה, לחץ/י 'סיימתי מותג זה'."
        ),
        reply_markup=build_model_keyboard(
            first_brand_idx,
            brand,
            sess.model_selection[first_brand_idx],
            page=0,
        ),
    )

# Page navigation inside the models list for a specific brand
async def model_page_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = _get_session(context)
    # data format: "model_page:<brand_idx>:<page>"
    _, bidx_str, page_str = query.data.split(":")
    bidx = int(bidx_str)
    page = int(page_str)
    sess.page_model[bidx] = page

    brand = sess.brands[bidx]
    selected = sess.model_selection.get(bidx, set())
    await query.edit_message_reply_markup(
        reply_markup=build_model_keyboard(bidx, brand, selected, page)
    )

# Toggle a model selection
async def model_toggle_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = _get_session(context)
    # data format: "model_toggle:<brand_idx>:<model_idx>:<page>"
    _, bidx_str, midx_str, page_str = query.data.split(":")
    bidx = int(bidx_str)
    midx = int(midx_str)
    page = int(page_str)

    sess.ensure_model_set(bidx)
    if midx in sess.model_selection[bidx]:
        sess.model_selection[bidx].remove(midx)
    else:
        sess.model_selection[bidx].add(midx)

    brand = sess.brands[bidx]
    await query.edit_message_reply_markup(
        reply_markup=build_model_keyboard(bidx, brand, sess.model_selection[bidx], page)
    )

# ✅ When user finishes selecting models for the *current* brand
async def models_done_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = _get_session(context)

    ordered = sorted(sess.brand_selection)
    if not ordered:
        await edit_text(query, "לא נבחרו מותגים. שלח/י /start.")
        return

    # איזה מותג סיימנו עכשיו?
    current_bidx = (
        int(sess.stage.split(":")[1]) if sess.stage.startswith("models:") else ordered[0]
    )

        # ✅ Require at least one model before continuing
    chosen_models = sess.model_selection.get(current_bidx, set())
    if not chosen_models:
        brand_name = sess.brands[current_bidx]['brand']
        await edit_text(
            query,
            f"לא נבחר אף דגם עבור {brand_name}. יש לבחור לפחות דגם אחד לפני שממשיכים.",
            reply_markup=build_model_keyboard(
                current_bidx,
                sess.brands[current_bidx],
                sess.model_selection[current_bidx],
                page=sess.page_model.get(current_bidx, 0)
            ),
        )
        return


    # ננסה לעבור למותג הבא שנבחר
    try:
        next_index = ordered.index(current_bidx) + 1
        next_bidx = ordered[next_index]
        sess.stage = f"models:{next_bidx}"
        sess.ensure_model_set(next_bidx)
        brand = sess.brands[next_bidx]

        await edit_text(
            query,
            text=(
                f"בחר/י דגמים עבור {brand['brand']}.\n"
                "כשתסיים/י עם המותג הזה, לחץ/י 'סיימתי מותג זה'."
            ),
            reply_markup=build_model_keyboard(
                next_bidx,
                brand,
                sess.model_selection[next_bidx],
                page=0,
            ),
        )
        return

    except (ValueError, IndexError):
        # אין עוד מותגים -> השלב הבא
        sess.stage = "ready"

    # אחרי שסיימנו לבחור הכל:
    # לפני הריצה אנחנו שואלים איזה סוג מודעות להחזיר (פרטי / סוכנות)
    await edit_text(
        query,
        text="בחר/י סוג מודעות להצגה (אפשר פרטי, סוכנות, או שניהם):",
        reply_markup=build_seller_keyboard(sess),
    )




def build_summary_text(sess: Session) -> str:
    lines = ["סיכום בחירה:"]
    for bidx in sorted(sess.brand_selection):
        brand = sess.brands[bidx]
        models = brand.get("models", [])
        model_names = [
            (models[midx].get("model_name") or str(models[midx].get("model_value")))
            for midx in sorted(sess.model_selection.get(bidx, []))
            if 0 <= midx < len(models)
        ]
        pretty = ', '.join(model_names) if model_names else 'No models selected (brand page only)'
        lines.append(f"• *{brand['brand']}* — {pretty}")

    # lines.append("לחץ/י Run כדי להתחיל סריקה. נקשר אליך עם תקציר + CSV בסיום.")
    # Seller filter line
    if sess.seller_filter == {"private", "agency"}:
        sel = "פרטי + סוכנות"
    elif sess.seller_filter == {"private"}:
        sel = "פרטי בלבד"
    elif sess.seller_filter == {"agency"}:
        sel = "סוכנות בלבד"
    else:
        sel = "—"
    lines.append(f"סוג מודעות: {sel}")
    lines.append("לחץ/י Run כדי להתחיל סריקה. נקשר אליך עם תקציר + CSV בסיום.")
    return "\n".join(lines)

async def go_scan_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.delete()
    # start the choosing menu (same logic you use in /scan)
    try:
        brands = json.loads(Path(JSON_FILENAME).read_text(encoding="utf-8"))
    except Exception as e:
        await query.message.reply_text(f"לא הצלחתי לטעון {JSON_FILENAME}: {e}")
        return

    context.user_data["session"] = Session(brands)
    await query.message.reply_text("בחר/י מותגים לסריקה (אפשר כמה):")
    await query.message.reply_text(
        "בחירת מותגים:",
        reply_markup=build_brand_keyboard(brands, set(), page=0),
    )

async def my_subs_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # delete the message this button was clicked from
    try:
        await query.message.delete()
    except Exception:
        pass

    tg_id = query.from_user.id
    subs = get_user_subscriptions(tg_id)

    if not subs:
        await query.message.reply_text("אין לך מנויים כרגע.")
        return

    # build the text
    lines = []
    for _, brand, model in subs:
        lines.append(f"- {brand} / {model if model.strip() else 'כל הדגמים'}")

    txt = (
        "המנויים הפעילים שלך:\n\n"
        + "\n".join(lines)
        + "\n\nאפשר להסיר מנוי ספציפי או למחוק הכל:"
    )

    # get the existing keyboard
    kb = build_subscriptions_keyboard(subs)

    # convert tuple → list so we can append
    try:
        rows = [list(r) for r in kb.inline_keyboard]   # InlineKeyboardMarkup
    except Exception:
        rows = [list(r) for r in kb]                   # fallback

    # add HOME button at the bottom
    rows.append([InlineKeyboardButton("🏠 תפריט ראשי", callback_data="home")])

    # rebuild markup
    kb_fixed = InlineKeyboardMarkup(rows)

    # send final message
    await query.message.reply_text(
        txt,
        reply_markup=kb_fixed,
        link_preview_options=LinkPreviewOptions(is_disabled=True),
    )


async def export_subs_csv_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    tg_id = query.from_user.id
    await _send_subs_csv_via(query, tg_id)

from openpyxl import Workbook
from openpyxl.styles import Alignment

async def _send_subs_csv_via(update_or_query, chat_id: int):
    """
    Export ONLY listings that match user's subscriptions,
    styled XLSX version (centered text, auto-fit columns, small padding).
    """
    subs = list_user_subscriptions(chat_id)
    if not subs:
        msg = "אין לך מנויים כרגע."
        if hasattr(update_or_query, "message") and update_or_query.message:
            await update_or_query.message.reply_text(msg)
        else:
            await update_or_query.reply_text(msg)
        return

    # Map subscriptions
    by_brand: dict[str, set[str]] = {}
    for s in subs:
        b = (s.get("brand") or "").strip()
        m = (s.get("model") or "").strip()
        if not b:
            continue
        by_brand.setdefault(b, set())
        if m:
            by_brand[b].add(m)
        else:
            by_brand[b] = set()

    all_rows, seen = [], set()
    def _row_key(r): return (r.get("url") or r.get("URL") or "").strip()

    for brand, models in by_brand.items():
        if models:
            for model in sorted(models):
                rows = query_selection([brand], [model], limit=100000)
                for r in rows:
                    k = _row_key(r)
                    if k not in seen:
                        seen.add(k)
                        all_rows.append(r)
        else:
            rows = query_selection([brand], [], limit=100000)
            for r in rows:
                k = _row_key(r)
                if k not in seen:
                    seen.add(k)
                    all_rows.append(r)

    if not all_rows:
        msg = "לא נמצאו מודעות תואמות למנויים שלך."
        if hasattr(update_or_query, "message") and update_or_query.message:
            await update_or_query.message.reply_text(msg)
        else:
            await update_or_query.reply_text(msg)
        return

    # Fixed schema (same order + Hebrew titles)
    headers = [
        ("url",           "קישור"),
        ("ad_created_at", "פורסם"),
        ("brand",         "יצרן"),
        ("model",         "דגם"),
        ("price",         "מחיר"),
        ("year",          "שנה"),
        ("km",            "ק״מ"),
        ("hands",         "ידיים"),
        ("location",      "מיקום"),
        ("seller_type",   "סוג מודעה"),
        ("first_seen_at", "נראה לראשונה"),
        ("last_seen_at",  "עודכן לאחרונה"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    # Write headers
    ws.append([h[1] for h in headers])

    # Write rows
    for r in all_rows:
        raw_url  = (r.get("url") or r.get("URL") or "").strip()
        created  = _format_date_only(r.get("ad_created_at") or "")
        brand    = (r.get("brand") or "").strip()
        model    = (r.get("model") or "").strip()
        price    = (r.get("price") or "").strip()
        year     = (r.get("year") or "").strip()
        km       = (r.get("km") or "").strip()
        hands    = (r.get("hands") or "").strip()
        location = (r.get("location") or "").strip()
        stype    = (r.get("seller_type") or "").strip().lower()
        first    = _format_date_only(r.get("first_seen_at") or "")
        last     = _format_date_only(r.get("last_seen_at") or "")

        # URL shortened clickable
        url_cell = ""
        if raw_url:
            url_cell = "קישור"
        if stype == "private":
            stype_he = "פרטי"
        elif stype == "agency":
            stype_he = "סוכנות"
        else:
            stype_he = stype

        ws.append([
            url_cell, created, brand, model, price, year, km,
            hands, location, stype_he, first, last
        ])

        # make last appended row's first cell a hyperlink
        if raw_url:
            ws.cell(row=ws.max_row, column=1).hyperlink = raw_url

    # Style: center alignment and small padding (auto column width)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            val = str(cell.value) if cell.value else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max_len + 4  # +4 padding

    filename = f"user_listings_{chat_id}.xlsx"
    wb.save(filename)

    try:
        if hasattr(update_or_query, "message") and update_or_query.message:
            await update_or_query.message.reply_document(
                document=open(filename, "rb"),
                filename=filename,
                caption="אקסל כל המנויים"
            )
        else:
            await update_or_query.reply_document(
                document=open(filename, "rb"),
                filename=filename,
                caption="אקסל כל המנויים"
            )
    except Exception as e:
        msg = f"שגיאה בשליחת הקובץ: {e}"
        # if hasattr(update_or_query, "message") and update_or_query.message:
        #     await update_or_query.message.reply_text(msg)
        # else:
        #     await update_or_query.reply_text(msg)
        if hasattr(update_or_query, "message") and update_or_query.message:
            await update_or_query.message.reply_text("פעולות נוספות:", reply_markup=_home_keyboard())
        else:
            await update_or_query.reply_text("פעולות נוספות:", reply_markup=_home_keyboard())


# -----------------------------
# Results helpers
# -----------------------------

def _apply_seller_filter(rows: list[dict], seller_filter: set[str] | None) -> list[dict]:
    if not seller_filter or seller_filter == {"private", "agency"}:
        return rows
    wanted = {x.lower() for x in seller_filter}
    def _stype(r: dict) -> str:
        return (r.get("seller_type") or "").strip().lower() or "private"
    return [r for r in rows if _stype(r) in wanted]


def build_seller_keyboard(sess: "Session") -> InlineKeyboardMarkup:
    priv_on = "✅ " if "private" in sess.seller_filter else ""
    ag_on   = "✅ " if "agency"  in sess.seller_filter else ""
    rows = [
        [InlineKeyboardButton(f"{priv_on}פרטי",  callback_data="seller_toggle:private"),
         InlineKeyboardButton(f"{ag_on}סוכנות", callback_data="seller_toggle:agency")],
        [InlineKeyboardButton("התחל סריקה ", callback_data="seller_done")],
        [InlineKeyboardButton("🏠 תפריט ראשי", callback_data="home")],
    ]
    return InlineKeyboardMarkup(rows)


#
async def seller_toggle_cb(update, context):
    query = update.callback_query
    await query.answer()
    sess = _get_session(context)
    _, which = query.data.split(":")
    which = which.strip().lower()  # "private" or "agency"

    # Toggle; never allow empty (must keep at least one checked)
    if which in sess.seller_filter:
        if len(sess.seller_filter) > 1:
            sess.seller_filter.remove(which)
    else:
        sess.seller_filter.add(which)

    await query.edit_message_reply_markup(reply_markup=build_seller_keyboard(sess))


async def seller_done_cb(update, context):
    query = update.callback_query
    await query.answer()

    # ✅ STEP 1 — quick reply so Telegram won't time out
    try:
        await query.edit_message_text("⏳ מעבד בקשה... זה יכול לקחת כמה שניות.")
    except Exception:
        pass
    await asyncio.sleep(0)

    # ✅ STEP 2 — OFFLOAD heavy work
    asyncio.create_task(_run_after_selection(query, context))

async def _run_after_selection(query, context):
    """Heavy path after user clicked 'start scan' – runs in background."""
    try:
        sess = _get_session(context)

        # stage guard
        if sess.stage != "ready":
            await query.message.chat.send_message(
                "עדיין לא סיימת את הבחירה. שלח/י /start כדי להתחיל מחדש."
            )
            return

        # required files exist?
        if not Path(SCRIPT_FILENAME).exists():
            await query.message.chat.send_message(f"לא נמצא הקובץ {SCRIPT_FILENAME} ליד tele_bot.py")
            return
        if not Path(JSON_FILENAME).exists():
            await query.message.chat.send_message(f"לא נמצא הקובץ {JSON_FILENAME} ליד tele_bot.py")
            return

        # user + plan
        tg_id = query.from_user.id
        username = (query.from_user.username or "") if query.from_user else ""
        upsert_user(tg_id, username)

        plan_name, MAX_SUBS = get_plan_limit(tg_id)
        MAX_SUBS = int(MAX_SUBS)

        # build (brand, model/None) pairs from selection and enforce limit
        pairs: list[tuple[str, str | None]] = []
        for bidx in sorted(sess.brand_selection):
            bname = sess.brands[bidx]["brand"]
            models = sess.brands[bidx].get("models", [])
            chosen = sorted(sess.model_selection.get(bidx, []))

            if chosen:
                for m in chosen:
                    if 0 <= m < len(models):
                        label = (models[m].get("model_name") or str(models[m].get("model_value")))
                        if count_user_subscriptions(tg_id) < MAX_SUBS:
                            pairs.append((bname, label))
                            add_subscription(tg_id, bname, label)
            else:
                if count_user_subscriptions(tg_id) < MAX_SUBS:
                    pairs.append((bname, None))
                    add_subscription(tg_id, bname, "")

        if count_user_subscriptions(tg_id) >= MAX_SUBS:
            await query.message.chat.send_message(f"הגעת למקסימום {MAX_SUBS} מנויים. לא הוספתי את כל הבחירות.")

        # enqueue / or send existing results for recent jobs
        enqueued = 0
        for (bname, mlabel) in pairs:
            bkey, mkey = _job_keys([bname], [mlabel] if mlabel else [])
            if recent_job_exists(tg_id, bkey, mkey, window_minutes=RUN_WINDOW_MINUTES):
                SKIP_LOG.info(
                    f"Skipped manual job for user={tg_id}, brand={bname}, model={mlabel or '(all)'} (last {RUN_WINDOW_MINUTES}m)."
                )
                # send current DB results instead of scraping
                await send_existing_results_to_user(
                    query,
                    [bname],
                    [mlabel] if mlabel else [],
                    max_items=10,
                    seller_filter=sess.seller_filter,
                )
                continue

            record_job_run(tg_id, bkey, mkey)
            await MANUAL_QUEUE.put(ScrapeJob(
                telegram_id=tg_id,
                brand_names=[bname],
                model_names=[mlabel] if mlabel else [],
                source="manual",
                seller_filter=list(sess.seller_filter),
            ))
            enqueued += 1

        # final user feedback + queue snapshot
        if enqueued == 0:
            await query.message.chat.send_message("לא הוספתי סריקה חדשה (כפילות בחלון הזמן האחרון).")
        else:
            await query.message.chat.send_message(f"התחלתי {enqueued} סריקות. אעדכן כאן עם התוצאות וה־XLSX.")

        print_queue_status(current=None)

    except Exception as e:
        # never crash silently
        try:
            await query.message.chat.send_message(f"שגיאה בהרצה: {e}")
        except Exception:
            pass



def _format_dt(dt_str: str) -> str:
    """
    Convert ISO 8601 (or close) datetime strings into readable Hebrew format.
    Example input: '2025-09-14T13:15:25Z'
    Output: '14/09/2025 16:15'
    """
    if not dt_str:
        return ""
    try:
        # Try parsing ISO8601 with or without Z
        dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
        # Format as DD/MM/YYYY HH:MM (24h clock)
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return dt_str  # fallback: return raw if parse fails




def _parse_date_only_or_none(s: str):
    if not s:
        return None

    s = s.strip()

    # dd/mm/yy
    m = re.match(r"(\d{2})/(\d{2})/(\d{2})", s)
    if m:
        d, m_, y = map(int, m.groups())
        y += 2000
        return datetime(y, m_, d).date()

    # ISO "2025-02-03T12:22:55Z"
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except:
        pass

    # Hebrew today/yesterday
    if s == "היום":
        return datetime.now().date()
    if s == "אתמול":
        return (datetime.now() - timedelta(days=1)).date()

    return None


def _new_banner_for_row(row: dict) -> str:
    """Choose the correct Hebrew banner for a new/boosted ad."""
    created_d = _parse_date_only_or_none((row.get("ad_created_at") or "").strip())
    first_seen_d = _parse_date_only_or_none((row.get("first_seen_at") or "").strip())
    today_d = datetime.now().date()  # server local date; OK for daily banner logic

    if created_d and created_d == today_d:
        return "🔥 מודעה חדשה 🔥"

    if created_d and first_seen_d and first_seen_d > created_d:
        return "⬆️ מודעה שהוקפצה ⬆️"

    # 3) Fallback
    return "🆕 מודעה חדשה"


def _home_keyboard():
    # Wide, non-breaking padding so Telegram won't collapse it:
    # \u00A0 = NO-BREAK SPACE, \u2007 = FIGURE SPACE (fixed-width-ish)
    PAD = ("\u00A0\u2007" * 8)  # tweak count to change width

    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"{PAD}🚀 התחלת סריקה{PAD}",   callback_data="go_scan")],
        [InlineKeyboardButton(f"{PAD}📬 המנויים שלי{PAD}",   callback_data="open_my_subs")],
        [InlineKeyboardButton(f"{PAD}📥 אקסל כל המנויים{PAD}",  callback_data="export_subs_csv")],
        # [InlineKeyboardButton(f"{PAD}💳 הרשמה למסלול{PAD}",  callback_data="register_plan")],
    ])


async def send_home_buttons_to_chat(app: Application, chat_id: int):
    await app.bot.send_message(
        chat_id=chat_id,
        text="פעולות נוספות:",
        reply_markup=_home_keyboard()
    )

def _main_menu_btn_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[InlineKeyboardButton("🏠 תפריט ראשי", callback_data="home")]])

async def home_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show the Home buttons (start-menu actions) on demand."""
    query = update.callback_query
    await query.answer()
    await query.message.delete()
    # Send a fresh message with the home keyboard (don’t edit long history)
    await query.message.chat.send_message(
        text="תפריט ראשי:",
        reply_markup=_home_keyboard(),
        link_preview_options=LinkPreviewOptions(is_disabled=True),
    )



def _rows_to_styled_xlsx(rows: list[dict], chat_id: int) -> str:
    """
    Build a styled XLSX from rows using the exact same schema/design as the
    Start-menu export: Hebrew headers, centered cells, auto-fit, short 'קישור'.
    Returns the filename saved.
    """
    # Fixed schema (order + Hebrew headers)
    headers = [
        ("url",           "קישור"),
        ("ad_created_at", "פורסם"),
        ("brand",         "יצרן"),
        ("model",         "דגם"),
        ("price",         "מחיר"),
        ("year",          "שנה"),
        ("km",            "ק״מ"),
        ("hands",         "ידיים"),
        ("location",      "מיקום"),
        ("seller_type",   "סוג מודעה"),
        ("first_seen_at", "נראה לראשונה"),
        ("last_seen_at",  "עודכן לאחרונה"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    # Header row
    ws.append([h[1] for h in headers])

    # Data rows
    for r in rows:
        raw_url  = (r.get("url") or r.get("URL") or r.get("link") or "").strip()
        created  = _format_date_only((r.get("ad_created_at") or "").strip())
        brand    = (r.get("brand") or r.get("Brand") or "").strip()
        model    = (r.get("model") or r.get("Model") or "").strip()
        price    = (r.get("price") or r.get("Price") or r.get("amount") or "").strip()
        year     = (r.get("year") or r.get("Year") or "").strip()
        km       = (r.get("km") or r.get("mileage") or r.get("Mileage") or "").strip()
        hands    = (r.get("hands") or "").strip()
        location = (r.get("location") or "").strip()
        stype    = (r.get("seller_type") or "").strip().lower()
        first    = _format_date_only((r.get("first_seen_at") or "").strip())
        last     = _format_date_only((r.get("last_seen_at") or "").strip())

        # Hebrew seller type
        if stype == "private":
            stype_he = "פרטי"
        elif stype == "agency":
            stype_he = "סוכנות"
        else:
            stype_he = stype

        # Short clickable URL label
        url_cell = "קישור" if raw_url else ""

        ws.append([
            url_cell, created, brand, model, price, year, km,
            hands, location, stype_he, first, last
        ])

        # Make the just-written row's URL a hyperlink
        if raw_url:
            ws.cell(row=ws.max_row, column=1).hyperlink = raw_url

    # Center & auto-fit with small padding
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max_len + 4  # minimal padding

    filename = f"user_listings_{chat_id}.xlsx"
    wb.save(filename)
    return filename


def _format_date_only(dt_str: str) -> str:
    """
    Return DD/MM/YYYY. Accepts ISO8601 variants (…Z) or raw; falls back to input.
    """
    if not dt_str:
        return ""
    try:
        dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y")
    except Exception:
        # try plain date like "2025-11-03"
        try:
            return datetime.strptime(dt_str[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            return dt_str


def _escape_md(s: str) -> str:
    """
    Basic Markdown escape to avoid breaking formatting.
    """
    if not s:
        return ""
    # Minimal escaping for Telegram Markdown (not V2) to keep it simple
    return (
        s.replace("*", r"\*")
         .replace("_", r"\_")
         .replace("[", r"\[")
         .replace("`", r"\`")
    )

def _shorten(text: str, max_len: int = 280) -> str:
    if not text:
        return ""
    t = text.strip()
    return t if len(t) <= max_len else (t[:max_len - 1].rstrip() + "…")

def _collect_specs(row: dict, max_items: int = 6) -> list[tuple[str, str]]:
    """
    Pulls a few useful 'extra' specs out of the fields JSON your scraper captured.
    We try to surface the most common ones; unknown keys are included up to a limit.
    """
    specs: list[tuple[str, str]] = []

    # Common top-level fields
    year  = (row.get("year") or row.get("Year") or "").strip()
    km    = (row.get("km") or row.get("mileage") or row.get("Mileage") or "").strip()
    hands = (row.get("hands") or "").strip()
    loc   = (row.get("location") or "").strip()

    if year:  specs.append(("📅 שנה", year))
    if loc:   specs.append(("📍 מיקום", loc))
    if km:    specs.append(("🛣️ ק״מ", km))
    if hands: specs.append(("🙋 ידיים", hands))


    # Extra ‘fields’ JSON (gearbox, trim, fuel, engine, color, owner type, etc.)
    fields_json = row.get("fields")
    try:
        fields = json.loads(fields_json) if isinstance(fields_json, str) else (fields_json or {})
    except Exception:
        fields = {}

    # Heuristic ordering – add the most useful car specs first if present (Heb/Eng)
    preferred_keys_order = [
        "תיבת הילוכים", "גיר", "Gearbox",
        "סוג דלק", "Fuel",
        "רמת גימור", "Trim",
        "נפח מנוע", "Engine", "Engine Capacity",
        "צבע", "Color",
        "בעלות", "Ownership",
        "מנוע", "Transmission",
    ]

    added = set()
    for k in preferred_keys_order:
        v = (fields.get(k) or "").strip()
        if v:
            specs.append((k, v))
            added.add(k)
            if len(specs) >= max_items:
                return specs

    # If room remains, add any other fields not yet shown
    for k, v in fields.items():
        if k in added:
            continue
        vv = (v or "").strip()
        if vv:
            specs.append((k, vv))
            if len(specs) >= max_items:
                break

    return specs

def format_listing_card(row: dict) -> str:
    """
    Rich, emoji-forward card with Markdown formatting.
    """
    # Title (fallback to brand + model)
    title = (row.get("title") or "").strip()
    if not title:
        b = (row.get("brand") or row.get("Brand") or "").strip()
        m = (row.get("model") or row.get("Model") or "").strip()
        title = (b + " " + m).strip()

    price = (row.get("price") or row.get("Price") or row.get("amount") or "").strip()
    url   = (row.get("url") or row.get("URL") or row.get("link") or "").strip()
    desc  = (row.get("description") or "").strip()

    first_seen = (row.get("first_seen_at") or "").strip()
    last_seen  = (row.get("last_seen_at") or "").strip()

    # Build the card
    lines: list[str] = []
    if title:
        lines.append(f"🚗 *{_escape_md(title)}*")
    if price:
        lines.append(f"💰 מחיר: {_escape_md(price)}")
    
    # Ad creation date (as scraped)
    created = (row.get("ad_created_at") or "").strip()
    if created:
        lines.append(f"📆 פורסם: {_escape_md(created)}")
    
    # Seller type (normalize to Hebrew)
    stype = (row.get("seller_type") or "").strip().lower()
    if stype:
        stype_he = "פרטי" if stype == "private" else "סוכנות"
        emoji = "🙋" if stype == "private" else "🏢"
        lines.append(f"{emoji} סוג מודעה: {stype_he}")
    



    # Key specs block
    specs = _collect_specs(row, max_items=8)
    for label, value in specs:
        # For labels already with emoji (our added ones), keep them;
        # for others, prepend a bullet to keep it uniform
        if label.startswith(("📅", "📍", "🛣️", "🙋")):
            lines.append(f"{label}: {_escape_md(value)}")
        else:
            lines.append(f"• {_escape_md(label)}: {_escape_md(value)}")

    # Description (short)
    if desc:
        lines.append("")
        lines.append(f"📝 תיאור:\n{_escape_md(_shorten(desc, 360))}")


    # Footer: timestamps (if available)
    footer_lines = []
    if first_seen:
        footer_lines.append(f"⌛ נראה לראשונה: {_format_dt(first_seen)}")
    if last_seen:
        footer_lines.append(f"⌛ עודכן לאחרונה: {_format_dt(last_seen)}")

    if footer_lines:
        lines.append("")
        lines.extend(_escape_md(x) for x in footer_lines)


    # Link at the end (kept, but not previewed)
    if url:
        lines.append("")
        lines.append(f"🔗 [לצפייה במודעה]({_escape_md(url)})")

    return "\n".join(lines).strip()


async def send_text(app, chat_id: int, text: str, **kwargs):
    return await app.bot.send_message(
        chat_id=chat_id,
        text=text,
        link_preview_options=LinkPreviewOptions(is_disabled=True),
        **kwargs
    )

async def edit_text(query, text: str, **kwargs):
    return await query.edit_message_text(
        text=text,
        link_preview_options=LinkPreviewOptions(is_disabled=True),
        **kwargs
    )


def _row_text(row: dict) -> str:
    # 1) Title (fallback to "brand model" if title missing)
    title = (row.get("title") or "").strip()
    if not title:
        b = (row.get("brand") or row.get("Brand") or "").strip()
        m = (row.get("model") or row.get("Model") or "").strip()
        title = (b + " " + m).strip()

    # 2) "price • year • km" (only include fields that exist)
    price = (row.get("price") or row.get("Price") or row.get("amount") or "").strip()
    year  = (row.get("year") or row.get("Year") or "").strip()
    km    = (row.get("km") or row.get("mileage") or row.get("Mileage") or "").strip()
    second = " • ".join([x for x in [price, year, km] if x])

    # 3) URL
    url = (row.get("url") or row.get("URL") or row.get("link") or "").strip()

    # Assemble EXACTLY like old bot: 3 lines, no markdown
    lines = [title]
    lines.append(second if second else "")
    lines.append(url if url else "")
    return "\n".join(lines).strip()



def _matches_selection(row: dict, sess: Session) -> bool:
    # Brand name must match one of selected brands (case-insensitive substring)
    brand_val = (row.get("brand") or row.get("Brand") or row.get("title") or "").lower()
    brand_names = [sess.brands[i]["brand"].lower() for i in sess.brand_selection]
    if not any(name in brand_val for name in brand_names):
        return False

    # If models chosen for that brand, require substring match in model/title fields
    titleish = (row.get("model") or row.get("Model") or row.get("title") or "").lower()
    for bidx in sess.brand_selection:
        models = sess.brands[bidx].get("models", [])
        chosen = sess.model_selection.get(bidx, set())
        if not chosen:
            # No model filter for this brand
            return True
        names = [
            (models[m].get("model_name") or str(models[m].get("model_value"))).lower()
            for m in chosen if 0 <= m < len(models)
        ]
        if any(n in titleish for n in names):
            return True
    return False


async def send_results_summary(query, sess: Session, max_items: int = 20):
    # Build name filters based on the user's selection (same as before)
    brand_names = [sess.brands[i]["brand"] for i in sorted(sess.brand_selection)]
    model_names = []
    for bidx in sorted(sess.brand_selection):
        models = sess.brands[bidx].get("models", [])
        chosen = sorted(sess.model_selection.get(bidx, []))
        if chosen:
            for m in chosen:
                if 0 <= m < len(models):
                    label = (models[m].get("model_name") 
                             or str(models[m].get("model_value")))
                    model_names.append(label)

    # Query rows from SQLite (unchanged source)
    # rows = query_selection(brand_names, model_names, limit=300)
    # rows = _apply_seller_filter(rows, seller_filter)
    rows = query_selection(brand_names, model_names, limit=300)
    rows = _apply_seller_filter(rows, sess.seller_filter)

    if not rows:
        await query.message.reply_text("לא נמצאו תוצאות תואמות לבחירה. מצרף את ה-CSV המלא.")
        # await try_send_csv(query)
        xlsx_path = _rows_to_styled_xlsx(rows, query.from_user.id)
        await query.message.reply_document(
            document=open(xlsx_path, "rb"),
            filename=xlsx_path,
            caption="תוצאות הסריקה (XLSX)"
        )
        try:
            os.remove(xlsx_path)
        except Exception:
            pass
        return

    # Send rich cards individually (nicest UX in Telegram)
    # Keep it reasonable to avoid rate limits; CSV will include the full set anyway.
    sent = 0
    for r in rows[:max_items]:
        card = format_listing_card(r)
        # Rich Markdown, no link preview (keep user in Telegram UI)
        await query.message.reply_text(
            card[:4000],
            parse_mode=ParseMode.MARKDOWN,
            link_preview_options=LinkPreviewOptions(is_disabled=True),
        )
        sent += 1

    # If there are more results, give a compact count line
    remaining = max(0, len(rows) - sent)
    if remaining > 0:
        await query.message.reply_text(
            f"…ויש עוד {remaining} תוצאות. מצרף CSV מלא.",
            link_preview_options=LinkPreviewOptions(is_disabled=True),
        )

    # Always attach the CSV for power users
    # await try_send_csv(query)
    
    xlsx_path = _rows_to_styled_xlsx(rows, query.from_user.id)
    await query.message.reply_document(
        document=open(xlsx_path, "rb"),
        filename=xlsx_path,
        caption="תוצאות הסריקה (XLSX)"
    )
    try:
        os.remove(xlsx_path)
    except Exception:
        pass

    
async def _debug_seed_queues():
    """
    Fill queues so aging will trigger:
      - 10 manual (to exceed ON threshold quickly)
      - 6 scheduled
    """
    def fake_job(src, i):
        return ScrapeJob(
            telegram_id=999000 + i,
            brand_names=[f"Brand{i%3}"],
            model_names=[f"Model{i%5}"],
            source=src,
            seller_filter=None,
        )

    for i in range(10):
        await MANUAL_QUEUE.put(fake_job("manual", i))

    for i in range(6):
        await SCHEDULED_QUEUE.put(fake_job("scheduler", i))


# -----------------------------
# App bootstrap
# -----------------------------
async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        textwrap.dedent(
            """
            *פקודות*
            /start — בחירת מותגים ודגמים והרצת הסורק
            /help — הצגת הודעת עזרה זו
            """

        ),
        parse_mode="Markdown",
    )
def _stdin_for_selection(brands_json: list[dict], brand_names: list[str], model_names: list[str]) -> str:
    """
    Convert brand/model NAMES to the exact stdin that yad2.py expects (1-based indices by JSON order).
    If model_names is empty for a brand => treat as "brand page only" (no models line).
    """
    # map names -> indices
    brand_indices = []
    models_for_brand: dict[int, list[int]] = {}
    for b_idx, b in enumerate(brands_json):
        if b["brand"] in brand_names:
            brand_indices.append(b_idx)
            # If user provided model_names, pick those models that match this brand
            if model_names:
                wanted = []
                for m_idx, m in enumerate(b.get("models", [])):
                    label = m.get("model_name") or str(m.get("model_value"))
                    if label in model_names:
                        wanted.append(m_idx)
                if wanted:
                    models_for_brand[b_idx] = wanted

    if not brand_indices:
        return "\n"

    first_line = ",".join(str(i+1) for i in brand_indices)
    lines = [first_line]
    for b_idx in brand_indices:
        mids = models_for_brand.get(b_idx, [])
        line = ",".join(str(m+1) for m in mids) if mids else ""
        lines.append(line)
    return "\n".join(lines) + "\n"

async def run_selection_once(app: Application, job: ScrapeJob):
    """
    Runs yad2.py for the selection (brands/models), waits, then diffs + notifies.
    """
    # Load JSON fresh (indices may change)
    brands = json.loads(Path(JSON_FILENAME).read_text(encoding="utf-8"))
    stdin_text = _stdin_for_selection(brands, job.brand_names, job.model_names)

    # Launch scraper
    proc = subprocess.Popen(
        ["python3", SCRIPT_FILENAME],
        stdin=subprocess.PIPE,
        stdout=None,
        stderr=None,
        text=True,
        bufsize=1,
        cwd=str(Path(__file__).resolve().parent),
    )
    assert proc.stdin is not None
    proc.stdin.write(stdin_text)
    proc.stdin.flush()
    proc.stdin.close()

    # Wait until the scraper exits (no timeout)
    ret = await asyncio.to_thread(proc.wait)
    if ret != 0:
        await send_text(app, job.telegram_id, f"⚠️ הסריקה נכשלה (קוד {ret}).")
        return

    # Query current rows from SQLite for this selection
    # rows = query_selection(job.brand_names, job.model_names, limit=500)
    rows = query_selection(job.brand_names, job.model_names, limit=500)
    try:
        rows = _apply_seller_filter(rows, set(job.seller_filter) if job.seller_filter else None)
    except Exception:
        pass

    # Compute diffs vs last time and update subscription_items
    new_listings, price_changes = diff_and_update_items(job.telegram_id, rows)
    print(f"[DEBUG] rows={len(rows)} new={len(new_listings)} price_changes={len(price_changes)} source={job.source}")

    if job.source == "manual":
        # Short summary only when there ARE changes; stay silent otherwise
        text_lines = []
        if new_listings:
            text_lines.append(f"✨ {len(new_listings)} מודעות חדשות.")
        if price_changes:
            text_lines.append(f"💸 {len(price_changes)} שינויי מחיר.")

        if text_lines:
            await send_text(app, job.telegram_id, "\n".join(text_lines))

        # Show up to 20 rich cards of THIS scan's rows (filtered above)
        sent = 0
        for r in rows[:20]:
            card = format_listing_card(r)
            await app.bot.send_message(
                chat_id=job.telegram_id,
                text=card[:4000],
                parse_mode=ParseMode.MARKDOWN,
                link_preview_options=LinkPreviewOptions(is_disabled=True),
            )
            sent += 1

        remaining = max(0, len(rows) - sent)
        if remaining > 0:
            await send_text(app, job.telegram_id, f"…ויש עוד {remaining} תוצאות.")

        # Send XLSX with the SAME styling as Start-menu export (only this scan)
        xlsx_path = _rows_to_styled_xlsx(rows, job.telegram_id)
        try:
            brand_str = " + ".join(job.brand_names) if job.brand_names else "—"
            model_str = " + ".join(job.model_names) if job.model_names else "כל הדגמים"
            await app.bot.send_document(
                chat_id=job.telegram_id,
                document=open(xlsx_path, "rb"),
                filename=xlsx_path,
                caption=f"תוצאות הסריקה — {brand_str} / {model_str}",
            )
            try:
                os.remove(xlsx_path)
            except Exception:
                pass

        except Exception as e:
            await send_text(app, job.telegram_id, f"שגיאה בשליחת הקובץ: {e}")

        # Show action buttons again so the user doesn't need to scroll
        await send_home_buttons_to_chat(app, job.telegram_id)
        
    else:
            # Scheduler run → send updates using the same rich cards (silent if none)
            sent_any = False
            
            # 1) New listings — use the dynamic banner helper too
            for row in new_listings[:20]:
                banner = _new_banner_for_row(row)
                card = banner + "\n\n" + format_listing_card(row)
                await app.bot.send_message(
                    chat_id=job.telegram_id,
                    text=card[:4000],
                    parse_mode=ParseMode.MARKDOWN,
                    link_preview_options=LinkPreviewOptions(is_disabled=True),
                )
                sent_any = True
            
            # 2) Price changes — unpack (row, old_price, new_price)
            for row, old_price, new_price in price_changes[:20]:
                # You can show a price-change banner explicitly:
                banner = f"💸 שינוי מחיר: {old_price} → {new_price}"
                # (Or keep using _new_banner_for_row(row) if you prefer the ‘new/reshare’ logic.)
                card = banner + "\n\n" + format_listing_card(row)
                await app.bot.send_message(
                    chat_id=job.telegram_id,
                    text=card[:4000],
                    parse_mode=ParseMode.MARKDOWN,
                    link_preview_options=LinkPreviewOptions(is_disabled=True),
                )
                sent_any = True
            


            subs = get_user_subscriptions(job.telegram_id)  # returns [(rowid, brand, model), ...]
            by_brand = defaultdict(set)
            for _, b, m in subs:
                b = (b or "").strip()
                m = (m or "").strip()
                if b in job.brand_names:
                    if m:
                        by_brand[b].add(m)   # specific model
                    else:
                        by_brand[b]          # brand-only (all models)

            # Update last_checked_at for exactly those subs (unchanged logic):
            for b, models in by_brand.items():
                if not models:
                    update_subscription_checked(job.telegram_id, b, "")
                else:
                    for m in models:
                        update_subscription_checked(job.telegram_id, b, m)



            # Show action buttons after scheduler notifications too (handy)
            if sent_any:
                await send_home_buttons_to_chat(app, job.telegram_id)



async def run_scraper_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = _get_session(context)

    if sess.stage != "ready":
        await edit_text(query, "עדיין לא סיימת את הבחירה. שלח/י /start כדי להתחיל מחדש.")
        return

    # Validate files exist
    if not Path(SCRIPT_FILENAME).exists():
        await edit_text(query, f"לא נמצא הקובץ {SCRIPT_FILENAME} ליד tele_bot.py")
        return
    if not Path(JSON_FILENAME).exists():
        await edit_text(query, f"לא נמצא הקובץ {JSON_FILENAME} ליד tele_bot.py")
        return

    # Persist the user
    tg_id = query.from_user.id
    username = (query.from_user.username or "") if query.from_user else ""
    upsert_user(tg_id, username)


    # Build per-brand model mapping straight from the session
    pairs: list[tuple[str, str|None]] = []
    brand_names = []
    for bidx in sorted(sess.brand_selection):
        bname = sess.brands[bidx]["brand"]
        brand_names.append(bname)
        models = sess.brands[bidx].get("models", [])
        chosen = sorted(sess.model_selection.get(bidx, []))
        if chosen:
            for m in chosen:
                if 0 <= m < len(models):
                    label = (models[m].get("model_name") or str(models[m].get("model_value")))
                    pairs.append((bname, label))
                    add_subscription(tg_id, bname, label)  # store exact pair
        else:
            pairs.append((bname, None))
            add_subscription(tg_id, bname, "")  # brand-only

    # Enqueue ONE job per pair (brand + [model])
    enqueued = 0
    for (bname, mlabel) in pairs:
        bkey, mkey = _job_keys([bname], [mlabel] if mlabel else [])
        if recent_job_exists(tg_id, bkey, mkey, window_minutes=RUN_WINDOW_MINUTES):
            SKIP_LOG.info(f"Skipped manual job for user={tg_id}, brand={bname}, model={mlabel or '(all)'} (last {RUN_WINDOW_MINUTES}m).")
            # Optional: show existing results for this specific pair
            await send_existing_results_to_user(
                query,
                [bname],
                [mlabel] if mlabel else [],
                max_items=10,
                seller_filter=sess.seller_filter,
            )
            continue

        record_job_run(tg_id, bkey, mkey)
        await MANUAL_QUEUE.put(ScrapeJob(
            telegram_id=tg_id,
            brand_names=[bname],
            model_names=[mlabel] if mlabel else [],
            source="manual",
            seller_filter=list(sess.seller_filter),
        ))


        enqueued += 1

    if enqueued == 0:
        await edit_text(query, f"לא הוספתי סריקה חדשה (כפילות ב-{RUN_WINDOW_MINUTES} דקות האחרונות)."
)
    else:
        await edit_text(query, f"התחלתי {enqueued} סריקות. אסיים ואז אשלח לך עדכון.")
    print_queue_status(current=None)



async def scheduler_loop(app: Application, interval_seconds: int = SCHEDULER_INTERVAL_SECONDS):
    """Every `interval_seconds`, enqueue a job per subscription (per user)."""
    while True:
        try:
            subs = list_subscriptions()
            # Group by user → collapse into brand/model name lists for each user
            by_user: dict[int, tuple[set[str], set[str]]] = {}
            for s in subs:
                uid = int(s["telegram_id"])
                b = (s["brand"] or "").strip()
                if not b:
                    continue

                m = (s["model"] or "").strip()
                if uid not in by_user:
                    by_user[uid] = (set(), set())
                by_user[uid][0].add(b)
                if m:
                    by_user[uid][1].add(m)

            for uid, (brands, models) in by_user.items():
                subs_rows = list_user_subscriptions(uid)
                for s in subs_rows:
                    b = (s.get("brand") or "").strip()
                    m = (s.get("model") or "").strip()
                    if not b:
                        continue
                    
                    # ✅ 1) same (brand,model,user) recently ran → skip (uses your 60m window)
                    bkey, mkey = _job_keys([b], [m] if m else [])
                    if recent_job_exists(uid, bkey, mkey, window_minutes=RUN_WINDOW_MINUTES):
                        SKIP_LOG.info(
                            f"Skipped scheduled job for user={uid}, brand={b}, model={m or '(all)'} "
                            f"(last {RUN_WINDOW_MINUTES}m)."
                        )
                        continue
                    
                    # ✅ 2) already pending in either queue → skip (short inline check, no new fn)
                    try:
                        _pending = list(SCHEDULED_QUEUE._queue) + list(MANUAL_QUEUE._queue)  # type: ignore[attr-defined]
                    except Exception:
                        _pending = []
                    if any(
                        getattr(j, "telegram_id", None) == uid and
                        _job_keys(getattr(j, "brand_names", []), getattr(j, "model_names", [])) == (bkey, mkey)
                        for j in _pending
                    ):
                        SKIP_LOG.info(
                            f"Skipped scheduled enqueue (pending exists) for user={uid}, "
                            f"brand={b}, model={m or '(all)'}."
                        )
                        continue
                    
                    # ✅ 3) record + enqueue
                    record_job_run(uid, bkey, mkey)
                    await SCHEDULED_QUEUE.put(ScrapeJob(
                        telegram_id=uid,
                        brand_names=[b],
                        model_names=[m] if m else [],
                        source="scheduler",
                        seller_filter=None,
                    ))

                  
        except Exception as e:
            # log to stdout; keep loop alive
            print("scheduler error:", e)
        await asyncio.sleep(interval_seconds)

# --- Queue to ensure only one scrape runs at a time ---
@dataclass
class ScrapeJob:
    telegram_id: int
    brand_names: list[str]
    model_names: list[str]
    source: str  # "manual" | "scheduler"
    seller_filter: list[str] | None = None

# We now keep two queues, so we can prioritize manual over scheduled
MANUAL_QUEUE: asyncio.Queue[ScrapeJob] = asyncio.Queue()
SCHEDULED_QUEUE: asyncio.Queue[ScrapeJob] = asyncio.Queue()

async def get_next_job():
    """
    Scheduled-dominant policy with light aging:
      - Normal: 3 scheduled, then 1 manual
      - Aging ON (manual backlog high): 2 scheduled, then 1 manual
      - Fall back to other queue if preferred is empty.
    """
    sched_w, manual_w, aging = _current_policy_weights()

    # if weights changed, reset streak so we don't carry old cadence
    if _POLICY_STREAK.get("weights") != (sched_w, manual_w):
        _POLICY_STREAK["kind"] = None
        _POLICY_STREAK["count"] = 0
        _POLICY_STREAK["weights"] = (sched_w, manual_w)

    kind = _POLICY_STREAK["kind"]
    count = _POLICY_STREAK["count"]

    # Decide which kind we want next based on streak vs weights
    if kind == "scheduled":
        prefer = "scheduled" if count < sched_w else "manual"
    elif kind == "manual":
        prefer = "manual"    if count < manual_w else "scheduled"
    else:
        # first pick of a cycle prefers scheduled (scheduled-dominant)
        prefer = "scheduled"

    # Try preferred queue first; fallback to the other if empty
    if prefer == "scheduled":
        if not SCHEDULED_QUEUE.empty():
            _POLICY_STREAK["kind"]  = "scheduled"
            _POLICY_STREAK["count"] = count + 1 if kind == "scheduled" else 1
            return await SCHEDULED_QUEUE.get()
        elif not MANUAL_QUEUE.empty():
            _POLICY_STREAK["kind"]  = "manual"
            _POLICY_STREAK["count"] = 1
            return await MANUAL_QUEUE.get()
    else:
        if not MANUAL_QUEUE.empty():
            _POLICY_STREAK["kind"]  = "manual"
            _POLICY_STREAK["count"] = count + 1 if kind == "manual" else 1
            return await MANUAL_QUEUE.get()
        elif not SCHEDULED_QUEUE.empty():
            _POLICY_STREAK["kind"]  = "scheduled"
            _POLICY_STREAK["count"] = 1
            return await SCHEDULED_QUEUE.get()

    # Nothing available right now
    await asyncio.sleep(0.5)
    return None

async def _free_trial_enforcer_loop(app: Application, days: int = 10, interval_hours: int = 24):
    """Once a day: find FREE users older than `days`, wipe their subs, and notify."""
    while True:
        try:
            expired = list_free_trial_expired_users(days=days)
            for tg_id, started_iso in expired:
                deleted = reset_user_subscriptions(tg_id)
                if deleted > 0:
                    try:
                        kb = InlineKeyboardMarkup([
                            # [InlineKeyboardButton("💳 הרשמה למסלול", callback_data="register_plan")],
                            [InlineKeyboardButton("🏠 תפריט ראשי",   callback_data="home")],
                        ])
                        await app.bot.send_message(
                            chat_id=tg_id,
                            text=(
                                "🔔 תקופת הניסיון בחינם (10 ימים) הסתיימה, לכן איפסתי את המנויים שלך ל-0.\n"
                                "כדי להמשיך לקבל התראות, בחר/י מסלול:"
                            ),
                            reply_markup=kb,
                            disable_web_page_preview=True,
                        )
                    except Exception:
                        pass
        except Exception:
            # don't crash the loop
            pass

        await asyncio.sleep(interval_hours * 3600)


async def scrape_worker(app: Application):
    """Background worker that dequeues scrape jobs with scheduled-dominant priority.

    Policy:
      - Prefer scheduled jobs (SCHEDULED_WEIGHT times in a row)
      - Then take a manual job (MANUAL_WEIGHT times)
      - Fall back to the other queue if the preferred one is empty
    """
    while True:
        # Choose the next job using the helper (scheduled-dominant)
        job = await get_next_job()
        if job is None:
            continue

        # Show current job in the dashboard
        current = _job_label(job)
        print_queue_status(current=current)

        try:
            if TEST_MODE:
                # Simulate fast dummy jobs (no real scraping)
                if job.source == "scheduler":
                    await asyncio.sleep(0.15)
                else:
                    await asyncio.sleep(0.35)
            else:
                # Normal scraping path
                await run_selection_once(app, job)

        except Exception as e:
            try:
                await send_text(app, job.telegram_id, f"⚠️ שגיאה בהרצה: {e}")
            except Exception:
                pass

        finally:
            # Mark correct queue as completed
            if job.source == "manual":
                MANUAL_QUEUE.task_done()
            else:
                SCHEDULED_QUEUE.task_done()

            print_queue_status(current=None)


def main():
    token = os.environ.get("TELEGRAM_TOKEN")
    if not token:
        raise SystemExit("Please set TELEGRAM_TOKEN env var.")

    # Ensure DB schemas exist
    init_subscriptions()
    init_job_runs()  # for 20-minute de-dup history
    init_plans() 
    # Start the background worker and scheduler once the bot is up
    async def on_start(_app: Application):
        asyncio.create_task(scrape_worker(_app))
        asyncio.create_task(scheduler_loop(_app, interval_seconds=3600))  # every 60 minutes
        # asyncio.create_task(scheduler_loop(_app, interval_seconds=900))  # every 15 minutes
        # asyncio.create_task(_free_trial_enforcer_loop(_app, days=10, interval_hours=24))
        if TEST_MODE:
            asyncio.create_task(_debug_seed_queues())


    defaults = Defaults(parse_mode=ParseMode.HTML)
    # app = Application.builder().token(token).defaults(defaults).build()
    # Build the application
    app = (
        Application
        .builder()
        .token(token)
        .defaults(defaults)
        .post_init(on_start)
        .build()
    )

    # Stripe (payment) handlers + webhook server
    register_stripe_handlers(app)
    boot_stripe_webhook_once()

    # Commands (non-payment)
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(home_cb, pattern=r"^home$"))
    app.add_handler(CommandHandler("reset_free_subs", reset_free_subs_cmd))
    app.add_handler(CallbackQueryHandler(reset_free_subs_cb, pattern=r"^reset_free_subs:(confirm|cancel)$"))



    app.add_handler(CommandHandler("scan", scan))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("my_subs", my_subs))
    app.add_handler(CommandHandler("reset_all_subs", reset_all_subs_cmd))

    # Subscription management
    app.add_handler(CallbackQueryHandler(unsubscribe_one_cb, pattern=r"^unsubscribe_one:\d+$"))
    app.add_handler(CallbackQueryHandler(unsubscribe_all_cb, pattern=r"^unsubscribe_all$"))
    app.add_handler(CallbackQueryHandler(export_subs_csv_cb, pattern=r"^export_subs_csv$"))
    app.add_handler(CommandHandler("myid", myid_cmd))
    app.add_handler(CommandHandler("set_plan", set_plan_cmd))
    app.add_handler(CommandHandler("get_plan", get_plan_cmd))



    # Selection pages (brands, models, seller filters)
    app.add_handler(CallbackQueryHandler(brand_page_cb,  pattern=r"^brand_page:"))
    app.add_handler(CallbackQueryHandler(brand_toggle_cb, pattern=r"^brand_toggle:"))
    app.add_handler(CallbackQueryHandler(brands_done_cb,  pattern=r"^brands_done$"))

    app.add_handler(CallbackQueryHandler(model_page_cb,  pattern=r"^model_page:"))
    app.add_handler(CallbackQueryHandler(model_toggle_cb, pattern=r"^model_toggle:"))
    app.add_handler(CallbackQueryHandler(models_done_cb,  pattern=r"^models_done:"))

    app.add_handler(CallbackQueryHandler(seller_toggle_cb, pattern=r"^seller_toggle:"))
    app.add_handler(CallbackQueryHandler(seller_done_cb,   pattern=r"^seller_done$"))

    # Running scraper
    app.add_handler(CallbackQueryHandler(run_scraper_cb, pattern=r"^run_scraper$"))
    app.add_handler(CallbackQueryHandler(go_scan_cb,      pattern=r"^go_scan$"))

    # Navigation
    app.add_handler(CallbackQueryHandler(my_subs_cb, pattern=r"^open_my_subs$"))

    # Start bot
    app.run_polling(close_loop=False)


if __name__ == "__main__":
    main()
